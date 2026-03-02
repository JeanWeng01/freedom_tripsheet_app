"""
Excel generation for Freedom trip sheets.
Loads the template .xlsx, fills in driver data cells, preserves all formulas.
"""

from pathlib import Path
from datetime import datetime
import shutil

import openpyxl

TEMPLATE_PATH = (
    Path(__file__).parent.parent
    / "Freedom Trip Sheet Template - SHARED WITH DRIVERS.xlsx"
)

OUTPUT_READY = Path(__file__).parent / "output" / "ready"
OUTPUT_FLAGGED = Path(__file__).parent / "output" / "flagged"

OUTPUT_READY.mkdir(parents=True, exist_ok=True)
OUTPUT_FLAGGED.mkdir(parents=True, exist_ok=True)

# Rows available for stop data in the template
STOP_ROW_START = 4
STOP_ROW_END = 20   # inclusive → 17 slots total
MAX_STOP_ROWS = STOP_ROW_END - STOP_ROW_START + 1  # 17

# MDC arriving/leaving → Excel label string
MDC_LABELS: dict[tuple[str, str], str] = {
    ("Bobtail", "SR trailer"):           "MDC - arrive with Bobtail, take SR trailer",
    ("Bobtail", "Linehaul trailer"):     "MDC - arrive with Bobtail, take LH trailer",
    ("SR trailer", "Bobtail"):           "MDC - drop SR trailer, leave with Bobtail",
    ("SR trailer", "SR trailer"):        "MDC - drop SR trailer, take trailer for another SR",
    ("Linehaul trailer", "Bobtail"):     "MDC - drop LH trailer, leave with Bobtail",
    ("Linehaul trailer", "SR trailer"):  "MDC - drop LH trailer, take SR trailer",
    ("Linehaul trailer", "Linehaul trailer"): "MDC - drop LH trailer, take LH trailer",
    # SR → Linehaul is handled separately (two rows)
}

TIME_FMT = "h:mm AM/PM"


def _minutes_to_fraction(minutes: int | None) -> float | None:
    """Convert minutes-since-midnight to Excel fractional day."""
    if minutes is None:
        return None
    return minutes / 1440.0


def _write_time(ws, cell_addr: str, minutes: int | None) -> None:
    cell = ws[cell_addr]
    if minutes is not None:
        cell.value = minutes / 1440.0
        # Only set format if the cell doesn't already have one from the template
        if not cell.number_format or cell.number_format == "General":
            cell.number_format = TIME_FMT
    else:
        cell.value = None


def _build_excel_rows(stops: list[dict]) -> list[dict]:
    """
    Convert the trip stop list into flat Excel row dicts.
    - Skips segment and truck stops (UI-only markers).
    - Skips stops that are skipped=True with no flag (no-order, no data).
    - Expands SR→Linehaul MDC into two rows.
    """
    rows: list[dict] = []

    for stop in stops:
        stype = stop.get("type", "")

        if stype in ("segment", "truck"):
            continue

        skipped = stop.get("skipped", False)
        flag = stop.get("flag")  # string or None

        # Truly skipped with no flag → omit from Excel entirely
        if skipped and not flag:
            continue

        arrival = _minutes_to_fraction(stop.get("arrivalTime"))
        departure = _minutes_to_fraction(stop.get("departureTime"))
        comment = stop.get("comment", "") or ""
        hub = stop.get("hubReading", "") or ""
        trailer = stop.get("trailerNumber", "") or ""
        reefer = stop.get("reeferTemp", "") or ""
        bol = stop.get("bolNumber", "") or ""

        if stype == "sr":
            location = stop.get("storeName") or stop.get("storeCode") or "Unknown store"
            rows.append({
                "flag": flag, "location": location,
                "arrival": arrival, "departure": departure,
                "comment": comment, "hub": hub,
                "trailer": trailer, "reefer": reefer, "bol": bol,
            })

        elif stype == "lh":
            location = stop.get("locationName", "") or "LH Location"
            rows.append({
                "flag": flag, "location": location,
                "arrival": arrival, "departure": departure,
                "comment": comment, "hub": hub,
                "trailer": trailer, "reefer": reefer, "bol": bol,
            })

        elif stype == "mdc":
            arriving = stop.get("arrivingWith")
            leaving = stop.get("leavingWith")
            special = stop.get("specialActivity")

            if special:
                # Special activity — single row with the activity label
                rows.append({
                    "flag": flag, "location": special,
                    "arrival": arrival, "departure": departure,
                    "comment": comment, "hub": hub,
                    "trailer": trailer, "reefer": reefer, "bol": bol,
                })

            elif arriving == "SR trailer" and leaving == "Linehaul trailer":
                # Split case: two rows
                transition = _minutes_to_fraction(stop.get("transitionTime"))
                rows.append({
                    "flag": flag, "location": "MDC - drop SR trailer",
                    "arrival": arrival, "departure": transition,
                    "comment": comment, "hub": hub,
                    "trailer": trailer, "reefer": reefer, "bol": bol,
                })
                rows.append({
                    "flag": None, "location": "MDC - take LH trailer",
                    "arrival": transition, "departure": departure,
                    "comment": "", "hub": "", "trailer": trailer,
                    "reefer": "", "bol": "",
                })

            else:
                key = (arriving, leaving) if arriving and leaving else None
                label = (
                    MDC_LABELS.get(key, f"MDC - {arriving} → {leaving}")
                    if key else "MDC Stop"
                )
                rows.append({
                    "flag": flag, "location": label,
                    "arrival": arrival, "departure": departure,
                    "comment": comment, "hub": hub,
                    "trailer": trailer, "reefer": reefer, "bol": bol,
                })

    return rows


def generate_excel(trip: dict) -> tuple[str, str, str]:
    """
    Generate a filled Excel file from trip data.

    Returns:
        (absolute_filepath, filename, status)
        status is 'ready' or 'flagged'
    """
    header = trip["header"]
    stops = trip.get("stops", [])

    # ── Filename ──────────────────────────────────────────────────────────────
    date_obj = datetime.strptime(header["date"], "%Y-%m-%d")
    date_str = date_obj.strftime("%m.%d.%Y")
    first_name = header["driverName"].split()[0]
    filename = f"[{date_str}] {header['routeNumber']} {first_name}.xlsx"

    # ── Load template ──────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["Trip Sheet"]

    # ── Header cells ──────────────────────────────────────────────────────────
    ws["A2"] = header["driverName"]
    ws["C2"] = header["routeNumber"]

    _write_time(ws, "D2", header.get("clockInTime"))

    # Finish time = departure of the last stop that has one
    active = [
        s for s in stops
        if s.get("type") not in ("segment", "truck")
        and not (s.get("skipped") and not s.get("flag"))
    ]
    finish_minutes: int | None = None
    for s in reversed(active):
        if s.get("departureTime") is not None:
            finish_minutes = s["departureTime"]
            break
    _write_time(ws, "E2", finish_minutes)

    # ── Build stop rows ────────────────────────────────────────────────────────
    excel_rows = _build_excel_rows(stops)

    status = "ready"
    error_reasons: list[str] = []

    if len(excel_rows) > MAX_STOP_ROWS:
        status = "flagged"
        error_reasons.append(
            f"Too many stops: {len(excel_rows)} rows needed, template has {MAX_STOP_ROWS} slots"
        )

    # Clear data columns in rows 4–20 (don't touch formula columns)
    DATA_COLS = ["A", "C", "D", "E", "F", "G", "H", "I", "J"]
    for row_num in range(STOP_ROW_START, STOP_ROW_END + 1):
        for col in DATA_COLS:
            ws[f"{col}{row_num}"].value = None

    # Write stop data
    for i, row in enumerate(excel_rows[:MAX_STOP_ROWS]):
        row_num = STOP_ROW_START + i

        if row["flag"]:
            ws[f"A{row_num}"] = row["flag"]

        ws[f"C{row_num}"] = row["location"]

        if row["arrival"] is not None:
            cell = ws[f"D{row_num}"]
            cell.value = row["arrival"]
            if not cell.number_format or cell.number_format == "General":
                cell.number_format = TIME_FMT

        if row["departure"] is not None:
            cell = ws[f"E{row_num}"]
            cell.value = row["departure"]
            if not cell.number_format or cell.number_format == "General":
                cell.number_format = TIME_FMT

        if row["comment"]:
            ws[f"F{row_num}"] = row["comment"]
        if row["hub"]:
            ws[f"G{row_num}"] = row["hub"]
        if row["trailer"]:
            ws[f"H{row_num}"] = row["trailer"]
        if row["reefer"]:
            ws[f"I{row_num}"] = row["reefer"]
        if row["bol"]:
            ws[f"J{row_num}"] = row["bol"]

    # ── Save ──────────────────────────────────────────────────────────────────
    out_path = (OUTPUT_READY if status == "ready" else OUTPUT_FLAGGED) / filename
    wb.save(out_path)

    # ── Basic verification (re-open, check cells we wrote) ────────────────────
    try:
        wb2 = openpyxl.load_workbook(str(out_path), data_only=True)
        ws2 = wb2["Trip Sheet"]

        if ws2["A2"].value != header["driverName"]:
            error_reasons.append(f"A2 driver name mismatch after save")

        if excel_rows and ws2["C4"].value is None:
            error_reasons.append("C4 first stop location is empty after save")

        if error_reasons and status == "ready":
            status = "flagged"
            flagged_path = OUTPUT_FLAGGED / filename
            shutil.move(str(out_path), str(flagged_path))
            out_path = flagged_path

    except Exception as e:
        error_reasons.append(f"Verification error: {e}")
        if status == "ready":
            status = "flagged"
            flagged_path = OUTPUT_FLAGGED / filename
            shutil.move(str(out_path), str(flagged_path))
            out_path = flagged_path

    return str(out_path), filename, status
